import matplotlib.pyplot as plt
import mysql.connector
import numpy as np
import os
import openpyxl

mydb = mysql.connector.connect(host = "localhost",
    user = "root",
    password = "ems@project12")




cursor = mydb.cursor()
cursor.execute("use emsproject")


# Year List
cursor.execute("select distinct BillYear from record")
years = cursor.fetchall()
years = [y for x in years for y in x]

if len(years) == 0:
    years.append(0)

print(years)


# Account ID list
cursor.execute("select AccountNo from accounts")
AccountIDs = cursor.fetchall()
AccountIDs = [y for x in AccountIDs for y in x]
if len(AccountIDs) == 0:
    AccountIDs.append(0)
print(AccountIDs)
months = ['January', 'February', 'March', 'April','May', 'June', 'July', 'August', 'September', 'October', 'November','December']
header = ['AccountID', 'Bill Year', 'Bill Month' ,'UnitsMonthly', 'KE Charges', 'Govt Charges', 'Total Bill']


def plot_graph_units(year,AccountNo):

    query  = f"""select r.BillMonth from record r join accounts a on a.AC_ID = r.AC_ID
                    where r.BillYear = {year} and a.AccountNo = '{AccountNo}' 
                   order by STR_TO_DATE(CONCAT('0001 ',BillMonth, ' 01'), '%Y %M %d') asc"""

    cursor.execute(query)
    month = cursor.fetchall()
    month = [y[:3] for x in month for y in x]
    # print(month)
    cursor.execute(f"""select r.UnitsMonthly from record r join accounts a on a.AC_ID = r.AC_ID
                    where r.BillYear = {year} and a.AccountNo = '{AccountNo}' 
                   order by STR_TO_DATE(CONCAT('0001 ',BillMonth, ' 01'), '%Y %M %d') asc""")
    units = cursor.fetchall()
    units = [y for x in units for y in x]
    # print(units)
    if units == None and len(units)!=0  :
        units =[]
        for x in range(len(month)):
            units.append(0)

    xaxis = np.array(month)
    yaxis = np.array(units)

    plt.title(f"Unit of {year}")
    plt.bar(xaxis,yaxis)
    plt.xlabel("Months")
    plt.ylabel("Units")
    # plt.show()



# plot_graph_units(2022,'0400009513165')





def plot_graph_price(year,AccountNo):
    cursor.execute(f"""select r.BillMonth from record r join accounts a on a.AC_ID = r.AC_ID
                    where r.BillYear = {year} and a.AccountNo = '{AccountNo}' 
                   order by STR_TO_DATE(CONCAT('0001 ',BillMonth, ' 01'), '%Y %M %d') asc""")
    month = cursor.fetchall()
    month = [y[:3] for x in month for y in x]

    cursor.execute(f"""select r.TotalBill from record r join accounts a on a.AC_ID = r.AC_ID
                    where r.BillYear = {year} and a.AccountNo = '{AccountNo}' 
                   order by STR_TO_DATE(CONCAT('0001 ',BillMonth, ' 01'), '%Y %M %d') asc"""
)
    amount = cursor.fetchall()
    amount = [y for x in amount for y in x]
    if amount == None or len(amount) <= 1:
        amount = []
        for x in range(len(month)):
            amount.append(0)


    # print(amount[0])
    xaxis = np.array(month)
    yaxis = np.array(amount)
    plt.title(f"Price of {year}")
    plt.plot(xaxis, yaxis)
    plt.xlabel("Months")
    plt.ylabel("Price")

    # plt.show()


# plot_graph_price(2023,'0400009513165')
column = ['AccountNo', 'Bill Month', 'Bill Year', 'Units Monthly', 'Govt Amount', 'KE Charges', 'Total Bill']
# print(column)


def add_data(data):

    header = []
    for val in range(len(column)):
        header.append(column[val].replace(" ",""))
    # print(header)
    header[0] ='AC_ID'

    entry = 0
    while entry<len(data):
        if data[entry] == "":
            data.pop(entry)
            header.pop(entry)
        elif entry>=2:
            data[entry] = float(data[entry])
            entry = entry+1
        else:
            entry = entry +1

    account_number = data[0]
    # # Check if Account is present in database or not
    # # if not present then append in the accountid list
    print(type(account_number))


    try:
        print(account_number)
        cursor.execute(f"Insert into accounts(AccountNo) value ('{account_number}')")
        mydb.commit()
        AccountIDs.append(account_number)
    except Exception as e:
        print(e)
    cursor.execute(f"select AC_ID from accounts where AccountNo = '{account_number}'")
    AC_ID = cursor.fetchall()
    # print(AC_ID[0][0])
    data[0] = AC_ID[0][0]



    insertquery = "Insert into record (" + ",".join(header) + f") value {tuple(data)}"

    cursor.execute(insertquery)
    mydb.commit()





# add_data(['0400009513009','June',2021,200,210.21,'',''])





def import_data(filename):

    excel_sheet = openpyxl.load_workbook(filename)
    # Activate the current active sheet
    ws = excel_sheet.active
    match = []
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Account Number':
            index_acc = col
        elif ws.cell(row=1, column=col).value == 'Bill Year':
            index_year = col
        match.append(ws.cell(row=1, column=col).value)

    match = [x.replace(" ",'') for x in match]
    for row in range(2, ws.max_row):
        account_number = ws.cell(row=row, column=index_acc).value
        if account_number not in AccountIDs:
            try:
                # print(account_number)
                cursor.execute(f"Insert into accounts(AccountNo) value ('{account_number}')")
                mydb.commit()
                AccountIDs.append(account_number)
            except Exception as e:
                print(e)

    match[index_acc-1] = 'AC_ID'



    for row in range(2, ws.max_row):
        data = []
        for col in range(1,ws.max_column+ 1):
            if col == index_acc:
                account_number = ws.cell(row=row, column=col).value
                # print(account_number)
                cursor.execute(f"select AC_ID from accounts where AccountNo = '{account_number}'")
                AC_ID = cursor.fetchone()[0]
                data.append(AC_ID)
            elif col == index_year:
                cell_year= ws.cell(row=row, column=col).value
                if cell_year not in years:
                    years.append(cell_year)
                data.append(cell_year)
            else:
                value = ws.cell(row=row, column=col).value
                data.append(value)
        insertquery = "Insert into record (" + ",".join(match) + f") values {tuple(data)}"
        cursor.execute(insertquery)
        mydb.commit()



# import_data("D:\\LabTaskPL\\Project\\Excel files\\refineDatanew.xlsx")


def export_data(dest_filenames):

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Energy Data"

    ws1.append(column)

    header = [x.replace(" ","") for x in column]
    print(header)

    query =  "select "+",".join(header) + " from record join accounts on record.AC_ID=accounts.AC_ID order by STR_TO_DATE(CONCAT(BillYear,BillMonth, ' 01'), '%Y %M %d')  asc"
    print(query)
    cursor.execute(query)
    data = cursor.fetchall()

    print(data)


#
    for row in data:
        ws1.append(list(row))


    wb.save(filename = dest_filenames)



# export_data("new2.xlsx")








